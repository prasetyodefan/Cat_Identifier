# ----------------------------------------------------------------
#?                       Read Pascal VOC                         |
# ----------------------------------------------------------------

import os
import time
start_pascal = time.time()

img_names = []
xml_names = []

for dirname, subdirs, filenames in os.walk('assets/Predict/np/'):
  for filename in filenames:
    if filename[-3:] != "xml":
      img_names.append(filename)
    else:
      xml_names.append(filename)

print("Total Files")
print("Images    :",len(img_names))
print("Xml Files :",len(xml_names))
print()

## Preprocessing step 2 - cropped images by bounding box using xml files 
import xmltodict
from matplotlib import pyplot as plt
from skimage.io import imread, imsave

path_annotations = "assets/Predict/np/"
path_images = "assets/Predict/np/"
class_names = ['bengal','persian','siamese','rblue','ragdoll']

images = []
target = []
gmb = []
cim = []
bnddata = []


def crop_bounding_box(img, bnd):
  x1, x2, y1, y2 = list(map(int, bnd.values()))
  _img = img.copy()
  _img = _img[y1:y2, x1:x2]
  _img = _img[:,:,:3]
  return _img

from PIL import Image, ImageDraw

for img_name in img_names:
  with open(path_annotations+img_name[:-4]+".xml") as fd:
    doc = xmltodict.parse(fd.read())

  img = imread(path_images+img_name)
  temp = doc["annotation"]["object"]
  if type(temp) == list:
    for i in range(len(temp)):
      if temp[i]["name"] not in class_names:
        continue
      bnddata.append(temp[i]["bndbox"])
      images.append(crop_bounding_box(img, temp[i]["bndbox"]))
      cim.append(crop_bounding_box(img, temp[i]["bndbox"]))
      gmb.append(img)
      target.append(temp[i]["name"])
  else:
    if temp["name"] not in class_names:
        continue
    bnddata.append(temp["bndbox"])
    images.append(crop_bounding_box(img, temp["bndbox"]))
    cim.append(crop_bounding_box(img, temp["bndbox"]))
    gmb.append(img)
    target.append(temp["name"])

# print total target by class
print("Total target by class")
for i in class_names:
  print(i , "   :", target.count(i))

# --------------------------------------------------------------------
end_pascal = time.time()
pascalex = end_pascal - start_pascal
print("Execution time Read Pascal : {} seconds".format(pascalex) ,"\n" )

# --------------------------------------------------------------------
# ----------------------------------------------------------------
#?                         PREPROCESS                            |
# ----------------------------------------------------------------
start_prepo = time.time()
## Preprocessing step 3 - resize images to 258x258 and normalize (remove background and grayscale)
import numpy as np
import skimage
from skimage.transform import resize

from scipy import ndimage
from skimage import io, color, exposure, filters
from skimage.filters import unsharp_mask

def resize_image(img, size = 32):
  _img = img.copy() 
  _img = resize(_img, (size, size))
  return _img

def remove_background(img):
  _img = img.copy()
  thresh = skimage.filters.threshold_otsu(_img)
  _img = (_img > thresh).astype(np.float32)
  return _img

def grayscale(img):
  _img = img.copy()
  _img = np.dot(_img[...,:3], [0.299, 0.587, 0.114])
  return _img

def prepo(img):
  _img = img.copy()

  # Apply Adaptive Histogram Equalization (AHE) to the grayscale image
  clahe_image = exposure.equalize_adapthist(_img , clip_limit=3)

  # Apply Unsharp Masking to the AHE result
  blurred = ndimage.gaussian_filter(clahe_image, sigma=1)
  unsharp_maskk = clahe_image - 0.5 * blurred

  # Perform segmentation using Thresholding
  threshold_value = unsharp_mask(unsharp_maskk, radius=5, amount=2)
  binary_image = unsharp_maskk > threshold_value

  # Convert the binary image to uint8 and scale it to 0-255
  binary_image = binary_image.astype(np.float32)
  return _img  


for i in range(len(images)):
  images[i] = resize_image(images[i])
  images[i] = grayscale(images[i])
  images[i] = prepo(images[i])
  # images[i] = remove_background(images[i])

## Extration step 1 - extract features using PHOG (Pyramid Histogram of Oriented Gradients)

# ----------------------------------------------------------------
#?                      EKSTRAKSI FITUR                          |
# ----------------------------------------------------------------
def calculate_lbp(img, radius=3, neighbors=8):
    
    lbp = np.zeros_like(img)
    
    for i in range(radius, img.shape[0] - radius):
        for j in range(radius, img.shape[1] - radius):
            center = img[i, j]
            binary_code = 0
            
            for k in range(neighbors):
                x = i + int(radius * np.cos(2 * np.pi * k / neighbors))
                y = j - int(radius * np.sin(2 * np.pi * k / neighbors))
                
                if img[x, y] >= center:
                    binary_code |= (1 << (neighbors - 1 - k))
            
            lbp[i, j] = binary_code
    
    return lbp


#--------------------------------------------------------------------------------

features = []
lbp1d = []
for i in range(len(images)):
    print("Processing", i+1, "of", len(images))
    lbp = calculate_lbp(images[i])  # Menghitung LBP untuk gambar
    lbp1d.append(lbp)
    # Mengubah dimensi LBP menjadi 1D
    lbp_1d = lbp.reshape(-1)
    
    features.append(lbp_1d)

# Mengubah dimensi fitur menjadi 2D
features_2d = np.array(features)
features_1d = np.array(lbp1d)

np.set_printoptions(threshold=np.inf)

featurearr = []

for i, feature in enumerate(features_1d):
  featurearr.append(feature)

# --------------------------------------------------------------------
end_prepo = time.time()
prepotime = end_prepo - start_prepo
print()
print("Execution time Prepp : {} seconds".format(prepotime))
# --------------------------------------------------------------------

import pickle

pkl_filename = 'app/Model/svm_model.pkl'
with open(pkl_filename, 'rb') as file:
    loaded_model = pickle.load(file)
    
# ----------------------------------------------------------------
#?                            PREDICT                            |
# ----------------------------------------------------------------
import random
leng = len(features)
# Create a DataFrame from your data
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Membuat DataFrame
data = []
kosong = 0


for p in range(leng):
    pdt = features_2d[p]
    ftr = featurearr[p]
    prediction = loaded_model.predict([pdt])
    data_test = target[p]
    im_name = img_names[p]
    # im_data = gmb[p]
    # imdata = Image.fromarray(im_data)
    prediction_str = str(prediction).strip("['']")
    data.append([p, im_name, data_test, prediction_str, kosong, ftr])

# ----------------------------------------------------------------
#?                          WRITE XLSX                           |
# ----------------------------------------------------------------
# df = pd.DataFrame(data, columns=['Index', 'Name', 'Data Test', 'Prediction', 'True', 'Image'])
df = pd.DataFrame(data, columns=['Index', 'Name', 'Data Test', 'Prediction', 'True', 'LBP'])

# Menyimpan DataFrame ke file Excel
output_filename = 'output.xlsx'
df.to_excel(output_filename, index=False)

# Membuka workbook yang sudah disimpan
workbook = load_workbook(output_filename)
sheet = workbook.active

# Mendapatkan huruf kolom untuk kolom ke-5 (True column)
column_letter = get_column_letter(5)

# Menambahkan rumus '=IF(C2=D2;1;0)' ke kolom True
for row in range(2, len(df) + 2):
    cell = f'{column_letter}{row}'
    sheet[cell].value = f'=IF(C{row}=D{row}, 1, 0)'

# Menambahkan rumus '=AVERAGE(E1:E51)' ke sel E5
cell = 'E252'
sheet[cell].value = '=AVERAGE(E1:E251)'
# Menyimpan workbook dengan rumus ke file Excel
workbook.save(output_filename)


print("Data exported to data.xlsx successfully.")


# ----------------------------------------------------------------
#?                          OVERLAY                              |
# ----------------------------------------------------------------
ovimg = []
def draw_bounding_box(image, bnddata, prediction, result_image_path):
    image_array = np.array(image)
    bndbox = x1, x2, y1, y2 = list(map(int, bnddata.values()))

    reshapebndbox = {
      
        'x1': bndbox[0],
        'x2': bndbox[1],
        'y1': bndbox[2],
        'y2': bndbox[3]
        
    }

    overlay_image_array = image_array.copy()
    overlay_image = Image.fromarray(overlay_image_array)

    draw = ImageDraw.Draw(overlay_image)
    draw.rectangle([(reshapebndbox['x1'], reshapebndbox['y1']), (reshapebndbox['x2'], reshapebndbox['y2'])], outline='red')

    text_position = (reshapebndbox['x1'] + 5, reshapebndbox['y1'] + 5)
    draw.text(text_position, prediction, fill='red')

    overlay_image.save(result_image_path)
    ovimg.append(overlay_image)

# Example usage
images = gmb
bnddata_list = bnddata
predictions = loaded_model.predict(features)
predictions_list = predictions.tolist()

for i, image in enumerate(images):
    bnddata = bnddata_list[i]
    prediction = predictions_list[i]
    result_image_path = f'assets/Pred_Result/Overlay/bengal{i+1}.jpg'  # Generate different names for each result image
    draw_bounding_box(image, bnddata, prediction, result_image_path)


# ----------------------------------------------------------------
#?                          DRAW PLOT                            |
# ----------------------------------------------------------------

import random
rnd = random.randint(0, len(features))

pdt = features[rnd]
im_name = img_names[rnd]
ovimgs = ovimg[rnd]
predt = loaded_model.predict([pdt])
prediction = str(predt).strip("['']") 

from sklearn.metrics import classification_report
from sklearn.metrics import hamming_loss,log_loss,hinge_loss,brier_score_loss
print('Len of Features : ', len(features))
print('Name : ', im_name)
print("Data Test  :", target[rnd])
print("Prediction :", prediction)
print(classification_report( target, predictions, target_names=class_names ))
print("hamming_loss : ",hamming_loss( target, predictions))
# print("log_loss : ",log_loss( target, predictions))
# print("hinge_loss : ",hinge_loss( target, predictions))
# print("brier_score_loss : ",brier_score_loss( target, predictions))
import matplotlib.pyplot as plt
from sklearn.metrics import ConfusionMatrixDisplay

# Create a figure and axes
cmd = ConfusionMatrixDisplay.from_predictions( target, predictions, display_labels=class_names)
fig, axes = plt.subplots(1, 3, figsize=(10, 5))

axes[0].imshow(gmb[rnd])
axes[0].set_title('Original')

axes[1].imshow(cim[rnd])
axes[1].set_title('Cropped ')

axes[2].imshow(ovimgs)
axes[2].set_title('Predict')


plt.tight_layout()
plt.show()